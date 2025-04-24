from mcp.server.fastmcp import FastMCP
from openpyxl import Workbook, load_workbook
import os
from pathlib import Path
from html import escape
from anthropic import Anthropic


anthropic = Anthropic()
mcp = FastMCP("mcp-excel-server")

global analysis_data


@mcp.tool()
def xlsx_to_html_table(
    file_path: str = "test.xlsx", sheet_name: str = "Sheet1", has_headers: bool = True
) -> str:
    """Use this tool when the user wants to analyse a file.
    The entire excel sheet is read into a html table format. You can use the same to
    do the analysis. The tool returns the html data"""

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]

        html_rows = []

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return f"Data at {file_path} is empty." + "<table></table>"

        if has_headers:
            header_row = rows[0]
            html_rows.append(
                "<tr>"
                + "".join(f"<th>{escape(str(col))}</th>" for col in header_row)
                + "</tr>"
            )
            data_rows = rows[1:]
        else:
            data_rows = rows

        for row in data_rows:
            html_rows.append(
                "<tr>"
                + "".join(
                    f"<td>{escape(str(cell)) if cell is not None else ''}</td>"
                    for cell in row
                )
                + "</tr>"
            )

        final_html = "<table>\n" + "\n".join(html_rows) + "\n</table>"

        return f"Use this Data in {file_path} file is for analysis :\n {final_html}"

    except Exception as e:
        return f"There was issue in reading the data: {e}"


@mcp.tool()
def create_excel_file(file_path: str = "test.xlsx", sheet_name: str = "Sheet1"):
    """Use this tool when a Excel file needs to be created.
    The location given has to be used as the file_path. You will
    receive confirmation on file creation."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(file_path)

        return (
            f"Excel file has been created at {file_path} with sheet_name {sheet_name}"
        )
    except Exception as e:
        return f"Faced exception: {e}. The task may not be completed."


# @mcp.tool()
def get_col_name(file_path: str = "test.xlsx", sheet_name: str = "Sheet1") -> str:
    """Use this tool when the user asks for the column names in the excel sheet.
    The row 1 is read from the excel sheet and returned in html format."""
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True))

        header_row = rows[0]
        columns = (
            "<tr>"
            + "".join(f"<th>{escape(str(col))}</th>" for col in header_row)
            + "</tr>"
        )
        return f"The columns for the {file_path} is {columns}"
    except Exception as e:
        return f"There was issue in reading the data: {e}"


def analyse_table(
    file_path: str = "test.xlsx", sheet_name: str = "Sheet1", has_headers: bool = True
) -> str:
    """Use this tool when the user specifically asks for an analysis of the given file.
    The entire excel sheet is read into a html table format, and then analysed by an expert analyst.
    Expect a feedback from the analyst on the report completion."""

    try:
        final_html = xlsx_to_html_table(file_path, sheet_name, has_headers)
        # sending the html to AI model for analysis

        hal_analysis_system = """You are hal4025, an expert in working on excel sheets.
                    You are very good in analysing xlsx files and providing the report.
                    Do not apologize. Do not provide guidance or examples. 
                    Do not share what you cannot do.
                    Please do not provide the python code for the user requests.
                    Users don't know to code.
                    Do not explain how analyis is being done."""

        messages = [{"role": "user", "content": final_html}]
        print(final_html)
        response = anthropic.messages.create(
            model="claude-3-5-haiku-20241022",
            max_tokens=1000,
            system=hal_analysis_system,
            messages=messages,
        )
        analysis_result = f"The analysis result {response.content[0].text}"

        return f"The analysis is ready. {analysis_result}."

    except Exception as e:
        return f"Unable to complete the analysis due to : {e}"


@mcp.tool()
def read_excel_file_cell(
    file_path: str = "test.xlsx",
    row_num: int = 1,
    col_num: int = 1,
    sheet_name: str = "Sheet1",
):
    """Use this tool when a particular cell in a excel file has to be read.
    Once data is read, expect the same to be returned"""
    try:
        wb = load_workbook(file_path, data_only=True)
        # data_only set to True will read the values and not the formulae
        ws = wb[sheet_name]
        wb.save(file_path)
        data = ws.cell(row=row_num, column=col_num).value
        return f"Data at row: {row_num} and col: {col_num} is {data}"
    except Exception as e:
        return f"There was issue in reading the data. {e}"


@mcp.tool()
def write_to_excel_file_cell(
    file_path: str = "test.xlsx",
    row_num: int = 1,
    col_num: int = 1,
    given: str = "9",
    sheet_name: str = "Sheet1",
):
    """Use this tool when data called given has to be written to a particular cell
    in a excel file that is provided. Once data is written, expect the confirmation to be returned"""
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        ws.cell(row=row_num, column=col_num).value = given
        wb.save(file_path)
        return f"Data {given} is written at row: {row_num} and col: {col_num}"
    except Exception as e:
        return f"There was issue in writing the data. {e}"


@mcp.tool()
def write_formula_to_excel_file_cell(
    file_path: str = "test.xlsx",
    row_num: int = 1,
    col_num: int = 3,
    formula: str = "=A1+B1",
    sheet_name: str = "Sheet1",
):
    """Use this tool when the given formula has to be written to a particular cell
    in a excel file that is provided. Once formula is written, expect the confirmation to be returned"""
    try:
        wb = load_workbook(file_path, data_only=False)
        ws = wb[sheet_name]
        ws.cell(row=row_num, column=col_num).value = formula
        wb.save(file_path)
        return f"Data {formula} is written at row: {row_num} and col: {col_num}"
    except Exception as e:
        return f"There was issue in writing the formula. {e}"


@mcp.tool()
def delete_value_at_cell(
    file_path: str = "test.xlsx",
    row_num: int = 1,
    col_num: int = 1,
    sheet_name: str = "Sheet1",
):
    """Use this tool when value at a particular cell
    in a excel file that is provided has to be deleted.
    Once data is written, expect the confirmation to be returned"""
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        ws.cell(row=row_num, column=col_num).value = None
        wb.save(file_path)
        return f"Data at row: {row_num} and col: {col_num} has been deleted"
    except Exception as e:
        return f"There was issue in deleting the data. {e}"


@mcp.tool()
def search_excel_sheet_for_value(
    file_path: str = "test.xlsx",
    sheet_name: str = "Sheet1",
    value_to_find: str = "5",
    match_case: bool = True,
):
    """Use this tool when a value_to_find has to to be located
    inside a excel file at the file_path"""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    results = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                val = str(cell.value)
                target = str(value_to_find)
                if match_case:
                    if val == target:
                        results.append((cell.row, cell.column))
                else:
                    if val.lower() == target.lower():
                        results.append((cell.row, cell.column))
    if len(results) > 0:
        return f"The location of the {value_to_find} is at {results}"  # List of (row, column) tuples
    else:
        return f"The {value_to_find} cannot be located in {file_path}"


@mcp.tool()
def write_analysis_md(
    file_path: str = "analysis_report.md",
):
    """Use this tool whenthe analysis report has to be written to the file.
    The data is taken from the analysis_data global variable, no need to look for data from outside.
    New report file will be generated every time, and you will get confirmation"""
    import uuid

    global analysis_data
    try:
        uid = uuid.uuid4()
        ver = str(uid)[:4]
        file_ver = ""

        if "." in file_path:
            filename = file_path.split(".")[0]
            file_ver = f"{filename}_{ver}.md"
        else:
            file_ver = f"{file_path}_{ver}.md"
        if analysis_data != "":
            with open(file_ver, "w") as anw:
                anw.write(analysis_data)

            # after writing analysis data, reset the global variable
            analysis_data = ""
            return f"The report has been written to {file_ver}"
        else:
            return "The analysis data is empty. Kindly get the analysis done."
    except Exception as e:
        return f"There was issue in writing report: {e}"


@mcp.tool()
def remove_file(file_path: str = "test.xlsx"):
    """Use this tool when a file needs to be deleted.
    The location given has to be used as file_path. You will
    get confirmation after the file deletion"""
    try:
        os.remove(file_path)
        return f"The file at {file_path} has been removed"
    except Exception as e:
        return f"Faced exception {e}. Task may not be completed."


if __name__ == "__main__":
    print("Starting the MCP Server")
    mcp.run()
